# 作者
# NATURE
# 日期
# 2022/12/21 16:23
# 功能
# 业务配置
from time import sleep
import xlrd2
import yaml
import xlwt
from openpyxl import load_workbook
from selenium.webdriver import ActionChains, Keys
from selenium.webdriver.common.by import By
from web_automation.waiqin.pageobject.login_page import LoginPage


class XtpzPage(LoginPage):
    def add_ywxf(self, ywlxs, gs='FCG', username='admin', password='123456'):
        '''
        :param ywlxs:业务类型
        :param gws: 管网类型
        :param gs: 公司
        :param username: 账号
        :param password: 登录密码
        :return:
        '''
        self.login(url,username, password)
        try:
            driver = self.get_driver()
            ActionChains(driver).move_to_element(driver.find_element(By.XPATH,"//span[@title='切换公司']")).perform()
            sleep(1)
            driver.find_element(By.XPATH,f"//div[@data-value='{gs}']").click()
        except Exception:
            pass
        xtpz = driver.find_elements(By.XPATH, '//span[@title="系统设置" and text()="系统设置"]')[0]
        # 将业务配置元素滑动到可见位置
        driver.execute_script('arguments[0].scrollIntoView()', xtpz)
        sleep(3)
        ActionChains(driver).move_to_element(xtpz).perform()
        sleep(3)
        elements = driver.find_elements(By.TAG_NAME, 'span')
        sleep(1)
        for element in elements:
            if element.text == '业务配置':
                print("点击业务配置")
                ActionChains(driver).click(element).perform()
                break
        sleep(2)
        print('进入业务配置')
        for ywlx, ywxfs in ywlxs.items():
            # 业务配置-查看
            driver.find_element(By.XPATH, r"//td[text()='" + ywlx + "']/../td[3]/div/button[1]").click()
            sleep(1)
            print("进入业务配置列表")
            for ywxf, gws in ywxfs.items():
                driver.find_element(By.XPATH, '//div[text()="+ 添加类型"]').click()
                sleep(2)
                if gs:
                    driver.find_element(By.ID, 'name').send_keys(gs + '-' + ywxf)
                    sleep(1)
                    driver.find_element(By.ID, 'code').send_keys(gs + '-' + ywxf)
                    sleep(1)
                else:
                    driver.find_element(By.ID, 'name').send_keys(ywxf)
                    sleep(1)
                    driver.find_element(By.ID, 'code').send_keys(ywxf)
                for gw in gws.keys():
                    if gw == '无':
                        continue
                    for sbxx in gws[gw]:
                        sb, zyls, sfdw, dwfs, dwbj, sffk, sfdwfk, fklc = sbxx
                        if sb == '无':
                            continue
                        # 添加作业对象
                        driver.find_element(By.XPATH, "//label[text()='作业对象']/../button[1]").click()
                        sleep(1)
                        # 点击设备类型
                        driver.find_element(By.XPATH, '//span[@class="ant-select-selection__rendered"]').click()
                        sleep(1)
                        try:
                            # 选择设备
                            driver.find_element(By.CSS_SELECTOR,
                                                "span[title='" + gw + "']+ul span[title='" + sb + "']").click()
                        except Exception as e:
                            print(f'------------------{ywxf}-{gw}下无{sb}设备-----------------------')
                            sleep(1)
                            driver.find_element(By.XPATH,
                                                "//span/div/button[@type='submit']/../button[@type='button']").click()
                            sleep(1)
                            continue
                        sleep(1)
                        # 作业历史选择
                        driver.find_element(By.XPATH,
                                            "//div[@id='historyRecords']/div/div[@class='ant-select-selection__rendered']").click()
                        sleep(1)
                        driver.find_element(By.XPATH, "//ul[ @role='listbox']/li[text()='" + zyls + "']").click()
                        sleep(1)
                        # 到位
                        if sfdw == '是':
                            driver.find_element(By.XPATH, "//span[text()='到位']").click()
                            sleep(1)
                            # 到位方式
                            driver.find_element(By.XPATH, "//span[text()='" + dwfs + "']").click()
                            # 到位半径
                            dwbj_ele = driver.find_element(By.XPATH, "//input[@role='spinbutton']")
                            dwbj_ele.send_keys(Keys.CONTROL, 'a')
                            sleep(1)
                            dwbj_ele.send_keys(dwbj)
                            sleep(1)
                        if sffk == '是':
                            # 是否反馈
                            driver.find_element(By.XPATH, "//span[text()='反馈']").click()
                            if sfdwfk == '是':
                                driver.find_element(By.XPATH, "//span[text()='到位后反馈']").click()
                            # 设备反馈
                            driver.find_element(By.XPATH,
                                                "//div[@id='equipFeedbackBpm']/div/div[@class='ant-select-selection__rendered']").click()
                            sleep(2)
                            # 设备流程
                            driver.find_element(By.XPATH, "//ul[@role='listbox']/li[text()='" + fklc + "']").click()
                            sleep(1)
                        # 保存
                        driver.find_element(By.XPATH, "//span/div/button[@type='submit']").click()
                        try:
                            n = 1
                            while driver.find_element(By.XPATH, "//span[text()='作业对象重复,请重新输入']"):
                                driver.find_element(By.XPATH, "//*[@id='objectName']").send_keys(n)
                                sleep(1)
                                n += 1
                                driver.find_element(By.XPATH, "//span/div/button[@type='submit']").click()
                                sleep(1)
                        except Exception as e:
                            pass
                        finally:
                            print(f'{ywlx}-{ywxf}-{gw}-{sb}已添加！')
                            # continue
                    print(f'{ywlx}-{ywxf}-{gw}所有设备已添加完毕！')
                # self.add_zydx(ywlx,ywxf,gws)
                # 点击-业务配置详情-保存
                driver.find_element(By.XPATH, '//span/button[1]').click()
                sleep(1)
            driver.find_element(By.XPATH, '//div[text()="业务配置"]').click()
            sleep(1)

    def add_zydx(self, ywlx, ywxf, gws):
        driver = self.get_driver()
        for gw in gws.keys():
            for sbxx in gws[gw]:
                sb, zyls, sfdw, dwfs, dwbj, sffk, sfdwfk, fklc = sbxx
                # 添加作业对象
                driver.find_element(By.XPATH, "//label[text()='作业对象']/../button[1]").click()
                sleep(1)
                # 点击设备类型
                driver.find_element(By.XPATH, '//span[@class="ant-select-selection__rendered"]').click()
                sleep(1)
                try:
                    # 选择设备
                    driver.find_element(By.CSS_SELECTOR, "span[title='" + gw + "']+ul span[title='" + sb + "']").click()
                except Exception as e:
                    print('------------------无' + gw + '-' + sb + '-----------------------')
                    sleep(1)
                    driver.find_element(By.XPATH, "//span/div/button[@type='submit']/../button[@type='button']").click()
                    sleep(1)
                    continue
                sleep(1)
                # 作业历史选择
                driver.find_element(By.XPATH,
                                    "//div[@id='historyRecords']/div/div[@class='ant-select-selection__rendered']").click()
                sleep(1)
                driver.find_element(By.XPATH, "//ul[ @role='listbox']/li[text()='" + zyls + "']").click()
                sleep(1)
                # 到位
                if sfdw == '是':
                    driver.find_element(By.XPATH, "//span[text()='到位']").click()
                    sleep(1)
                    # 到位方式
                    driver.find_element(By.XPATH, "//span[text()='" + dwfs + "']").click()
                    # 到位半径
                    dwbj_ele = driver.find_element(By.XPATH, "//input[@role='spinbutton']")
                    dwbj_ele.send_keys(Keys.CONTROL, 'a')
                    sleep(1)
                    dwbj_ele.send_keys(dwbj)
                    sleep(1)
                if sffk == '是':
                    # 是否反馈
                    driver.find_element(By.XPATH, "//span[text()='反馈']").click()
                    if sfdwfk == '是':
                        driver.find_element(By.XPATH, "//span[text()='到位后反馈']").click()
                    # 设备反馈
                    driver.find_element(By.XPATH,
                                        "//div[@id='equipFeedbackBpm']/div/div[@class='ant-select-selection__rendered']").click()
                    sleep(2)
                    # 设备流程
                    driver.find_element(By.XPATH, "//ul[@role='listbox']/li[text()='" + fklc + "']").click()
                    sleep(1)
                # 保存
                driver.find_element(By.XPATH, "//span/div/button[@type='submit']").click()
                try:
                    n = 1
                    while driver.find_element(By.XPATH, "//span[text()='作业对象重复,请重新输入']"):
                        driver.find_element(By.XPATH, "//*[@id='objectName']").send_keys(n)
                        sleep(1)
                        n += 1
                        driver.find_element(By.XPATH, "//span/div/button[@type='submit']").click()
                        sleep(1)
                except Exception as e:
                    pass
                finally:
                    print(f'{ywlx}-{ywxf}{gw}-{sb}已添加！')
                    # continue
            print(f'{ywlx} -{ywxf}-{gw} 所有设备已添加完毕！')

    @staticmethod
    def get_null_cell(file):
        '''读取sheet页中空白单元格'''
        s = xlrd2.open_workbook(file).sheet_by_index(0)
        k = []
        for row in range(1, s.nrows):
            for col in range(0, s.ncols):
                if not s.cell_value(row, col):
                    k.append((row, col))
        return k

    @staticmethod
    def write_null_cell(file):
        '''填充sheet页中空白单元格'''
        k = XtpzPage.get_null_cell()
        # 加载excel，注意路径要与脚本一致
        wb = load_workbook(file)
        # 激活excel表
        sheet = wb.active
        for lc in k:
            row, col = lc
            try:
                sheet.cell(row + 1, col + 1, '无')
            except Exception:
                pass
        wb.save(file)
        wb.close()

    @staticmethod
    def get_ywpz_info_by_xlsx(file):
        '''从Excel中获取业务配置信息'''
        # XtpzPage.write_null_cell()
        # 获取空单元格
        s = xlrd2.open_workbook(file).sheet_by_index(0)
        k = []
        for row in range(1, s.nrows):
            for col in range(0, s.ncols):
                if not s.cell_value(row, col):#获取值从0开始
                    k.append((row, col))

        # 给空单元格赋值
        # 加载excel，注意路径要与脚本一致
        wb = load_workbook(file)
        # 激活excel表
        sheet = wb.active
        for lc in k:
            row, col = lc
            try:
                sheet.cell(row + 1, col + 1, '无')#赋值从1开始
            except Exception:
                pass
        wb.save(file)
        wb.close()

        # 获取业务类型数据
        s = xlrd2.open_workbook(file).sheet_by_index(0)
        ywlxs = {}
        #业务类型、业务细分、管网合并单元格值为空处理
        for row in range(1, s.nrows):
            ywlx, ywxf, gw, sb, zyls, sfdw, dwfs, dwbj, sffk, sfdwfk, fklc = s.row_values(row)
            n = row
            while n > 1:
                if ywlx:
                    break
                else:
                    ywlx = s.cell_value(n - 1, 0)
                    # print(ywlx)
                    n = n - 1
            n = row
            while n > 1:
                if ywxf:
                    break
                else:
                    ywxf = s.cell_value(n - 1, 1)
                    # print(ywlx)
                    n = n - 1
            n = row
            while n > 1:
                if gw:
                    break
                else:
                    gw = s.cell_value(n - 1, 2)
                    # print(ywlx)
                    n = n - 1
            if not sb:
                sb = '无'
            print(ywlx, ywxf, gw, sb, zyls, sfdw, dwfs, dwbj, sffk, sfdwfk, fklc)
            if ywlx in ywlxs.keys():
                if ywxf in ywlxs[ywlx].keys():
                    if gw in ywlxs[ywlx][ywxf].keys():
                        ywlxs[ywlx][ywxf][gw].append([sb, zyls, sfdw, dwfs, dwbj, sffk, sfdwfk, fklc])
                    else:
                        ywlxs[ywlx][ywxf][gw] = [[sb, zyls, sfdw, dwfs, dwbj, sffk, sfdwfk, fklc]]
                else:
                    ywlxs[ywlx][ywxf] = {gw: [[sb, zyls, sfdw, dwfs, dwbj, sffk, sfdwfk, fklc]]}
            else:
                ywlxs[ywlx] = {ywxf: {gw: [[sb, zyls, sfdw, dwfs, dwbj, sffk, sfdwfk, fklc]]}}
        return ywlxs

    @staticmethod
    def get_ywpz_info_by_yaml():
        with open('../data/xtpz.yaml', 'r', encoding="utf-8") as f:
            ywlxs = yaml.load(f, Loader=yaml.FullLoader)
        return ywlxs


if __name__ == '__main__':
    url='http://10.41.16.20:32091/?ecode=FCG&init=FCG'
    username='admin'
    password='Gis@123456'
    file='../data/ywpz_info.xlsx'
    ywlxs = XtpzPage.get_ywpz_info_by_xlsx(file)
    print(ywlxs)
    x = XtpzPage()
    x.add_ywxf(ywlxs,username=username,password=password)
