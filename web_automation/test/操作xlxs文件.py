# 作者
# NATURE
# 日期
# 2022/10/14 17:52
# 功能
#
import os

import openpyxl
# 创建一个Workbook对象
wb = openpyxl.Workbook()
# 如果不指定sheet索引和表名，默认在第二张表位置新建表名sheet1
wb.create_sheet(index=1, title="sheet2")

# 获取当前活跃的sheet，默认为第一张sheet
ws = wb.active
print(ws)
# 获取当前工作簿的所有sheet对象
sheets = wb.worksheets
print(sheets)
# 获取所有sheet的名字
sheets_name = wb.sheetnames
print(sheets_name)

# 保存为工作簿data1.xlsx
wb.save(os.path.dirname(__file__).split('test')[0]+'commons\data.xlsx')
