import openpyxl as openpyxl
# 读取数据
excel=openpyxl.load_workbook('../datas/test.xlsx')
print(excel.sheetnames)#获取所有sheet页
sheet=excel['Sheet1']
for row in excel['Sheet1'].values:#以行读取所有单元格数据存入元组中，没有值以none填充
    print(row)
# 获取单元格数据
print(sheet['A1'].value)
print(sheet.cell(1,1).value)

